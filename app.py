from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import json
import os
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

# Import our secure modules
from config import config
from database import db, DatabaseManager
from auth import AuthManager, login_required, role_required

# Create Flask app with configuration
app = Flask(__name__)
config_name = os.getenv('FLASK_ENV', 'development')
app.config.from_object(config[config_name])

# Initialize rate limiter
limiter = Limiter(
    app,
    key_func=get_remote_address,
    default_limits=["100 per hour"]
)

# Create necessary directories
os.makedirs(app.config['UPLOAD_PATH'], exist_ok=True)
os.makedirs('data', exist_ok=True)
os.makedirs('backups', exist_ok=True)

# Initialize database
db_manager = DatabaseManager()

# Migrate existing JSON data if it exists
if os.path.exists('raffle_data.json'):
    db_manager.migrate_from_json('raffle_data.json')

# Security middleware
@app.before_request
def security_headers():
    # Add security headers
    pass

@app.after_request
def after_request(response):
    # Add security headers
    for header, value in app.config['SECURITY_HEADERS'].items():
        response.headers[header] = value
    return response

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def process_excel_file(filepath):
    """
    Process Excel file and extract employee names using openpyxl.
    This function will try to find employee names in common column patterns.
    """
    try:
        # Load the Excel file
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        # Get all data from the sheet
        data = []
        headers = []
        
        # Read headers (first row)
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value else "")
        
        # Read all data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append([str(cell) if cell is not None else "" for cell in row])
        
        # Common column name patterns that might contain employee names
        name_patterns = ['name', 'caregiver', 'employee', 'staff']
        
        employees = []
        name_col_index = None
        detected_name_column = None
        
        # Find the name column by checking headers
        for i, header in enumerate(headers):
            if header and any(pattern in header.lower() for pattern in name_patterns):
                name_col_index = i
                detected_name_column = header
                break
        
        if name_col_index is not None:
            # Extract names from the identified column
            for row in data:
                if len(row) > name_col_index and row[name_col_index]:
                    name = str(row[name_col_index]).strip()
                    if (len(name) > 2 and len(name) < 50 and 
                        any(c.isalpha() for c in name) and 
                        name.lower() not in ['none', 'null', 'nan', '']):
                        employees.append(name)
        else:
            # If no obvious name column, check first few columns for potential names
            for col_index in range(min(3, len(headers))):  # Check first 3 columns
                for row in data:
                    if len(row) > col_index and row[col_index]:
                        potential_name = str(row[col_index]).strip()
                        if (len(potential_name) > 2 and len(potential_name) < 50 and 
                            any(c.isalpha() for c in potential_name) and
                            potential_name.lower() not in ['none', 'null', 'nan', '']):
                            employees.append(potential_name)
        
        if workbook:
            workbook.close()
        
        return {
            'success': True,
            'employees': list(set(employees)),  # Remove duplicates
            'total_rows': len(data),
            'columns': headers,
            'detected_name_column': detected_name_column
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    if request.method == 'POST':
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        success, message, user_data = AuthManager.login(
            email, password, get_remote_address()
        )
        
        if success:
            token = AuthManager.generate_token(user_data)
            session['access_token'] = token
            session['user_id'] = user_data['id']
            session['user_role'] = user_data['role']
            
            return jsonify({
                'success': True,
                'message': message,
                'token': token,
                'user': user_data
            })
        else:
            return jsonify({'success': False, 'message': message}), 401
    
    # Check if user is already logged in
    if 'access_token' in session:
        token_data = AuthManager.verify_token(session['access_token'])
        if token_data:
            return redirect(url_for('dashboard'))
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    # Log the logout
    if 'user_id' in session:
        db.log_audit(session['user_id'], "User logout", ip_address=get_remote_address())
    
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/api/employees', methods=['GET'])
@login_required
def get_employees():
    try:
        with db.get_connection() as conn:
            cursor = conn.execute('''
                SELECT id, name, email, phone, department, position, 
                       hire_date, photo_path, total_entries, is_active,
                       created_at, updated_at
                FROM employees WHERE is_active = 1
                ORDER BY name
            ''')
            
            employees = []
            for row in cursor.fetchall():
                employee = dict(row)
                # Get recent activities
                activity_cursor = conn.execute('''
                    SELECT activity_name, activity_category, entries_awarded, created_at
                    FROM activities WHERE employee_id = ?
                    ORDER BY created_at DESC LIMIT 10
                ''', (employee['id'],))
                
                employee['activities'] = [dict(activity) for activity in activity_cursor.fetchall()]
                employees.append(employee)
            
            return jsonify({'success': True, 'employees': employees})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee', methods=['POST'])
@login_required
@role_required('manager')
def add_employee():
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        email = data.get('email', '').strip().lower()
        phone = data.get('phone', '').strip()
        department = data.get('department', '').strip()
        position = data.get('position', '').strip()
        hire_date = data.get('hire_date')
        
        if not name:
            return jsonify({'success': False, 'error': 'Employee name is required'}), 400
        
        if email and not AuthManager.validate_email(email):
            return jsonify({'success': False, 'error': 'Invalid email format'}), 400
        
        with db.get_connection() as conn:
            # Check if employee already exists
            cursor = conn.execute('SELECT id FROM employees WHERE name = ? OR (email = ? AND email != "")', (name, email))
            if cursor.fetchone():
                return jsonify({'success': False, 'error': 'Employee already exists'}), 400
            
            # Insert new employee
            cursor = conn.execute('''
                INSERT INTO employees (name, email, phone, department, position, hire_date)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (name, email or None, phone or None, department or None, position or None, hire_date))
            
            employee_id = cursor.lastrowid
            conn.commit()
            
            # Log the action
            db.log_audit(
                request.current_user['user_id'],
                "Added employee",
                "employees",
                employee_id,
                new_values=data,
                ip_address=get_remote_address()
            )
            
            return jsonify({'success': True, 'employee_id': employee_id})
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee/<int:employee_id>/add_entry', methods=['POST'])
@login_required
@role_required('manager')
def add_entry(employee_id):
    try:
        data = request.get_json()
        activity_name = data.get('activity_name', '').strip()
        activity_category = data.get('activity_category', '').strip()
        entries_awarded = int(data.get('entries_awarded', 1))
        notes = data.get('notes', '').strip()
        
        if not activity_name:
            return jsonify({'success': False, 'error': 'Activity name is required'}), 400
        
        if entries_awarded <= 0 or entries_awarded > 10:
            return jsonify({'success': False, 'error': 'Entries must be between 1 and 10'}), 400
        
        with db.get_connection() as conn:
            # Check if employee exists
            cursor = conn.execute('SELECT id, name, total_entries FROM employees WHERE id = ? AND is_active = 1', (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                return jsonify({'success': False, 'error': 'Employee not found'}), 404
            
            # Add activity
            cursor = conn.execute('''
                INSERT INTO activities (employee_id, activity_name, activity_category, 
                                      entries_awarded, awarded_by, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (employee_id, activity_name, activity_category, entries_awarded, 
                 request.current_user['user_id'], notes))
            
            # Update employee total entries
            new_total = employee['total_entries'] + entries_awarded
            conn.execute('UPDATE employees SET total_entries = ? WHERE id = ?', 
                        (new_total, employee_id))
            
            conn.commit()
            
            # Log the action
            db.log_audit(
                request.current_user['user_id'],
                f"Added {entries_awarded} raffle entries",
                "activities",
                cursor.lastrowid,
                new_values={
                    'employee_name': employee['name'],
                    'activity': activity_name,
                    'entries': entries_awarded
                },
                ip_address=get_remote_address()
            )
            
            return jsonify({
                'success': True, 
                'message': f'Added {entries_awarded} entries for {activity_name}',
                'new_total': new_total
            })
            
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid entries value'}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee/<int:employee_id>', methods=['DELETE'])
@login_required
@role_required('admin')
def delete_employee(employee_id):
    try:
        with db.get_connection() as conn:
            # Get employee data before deletion
            cursor = conn.execute('SELECT name FROM employees WHERE id = ?', (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                return jsonify({'success': False, 'error': 'Employee not found'}), 404
            
            # Soft delete - mark as inactive
            conn.execute('UPDATE employees SET is_active = 0 WHERE id = ?', (employee_id,))
            conn.commit()
            
            # Log the action
            db.log_audit(
                request.current_user['user_id'],
                "Deleted employee (soft delete)",
                "employees",
                employee_id,
                old_values={'name': employee['name']},
                ip_address=get_remote_address()
            )
            
            return jsonify({
                'success': True, 
                'message': f'Employee {employee["name"]} has been removed'
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/employee/<int:employee_id>/reset_points', methods=['POST'])
@login_required
@role_required('admin')
def reset_employee_points(employee_id):
    try:
        with db.get_connection() as conn:
            # Get employee data
            cursor = conn.execute('SELECT name, total_entries FROM employees WHERE id = ? AND is_active = 1', (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                return jsonify({'success': False, 'error': 'Employee not found'}), 404
            
            old_total = employee['total_entries']
            
            # Reset points
            conn.execute('UPDATE employees SET total_entries = 0 WHERE id = ?', (employee_id,))
            
            # Add activity record for the reset
            conn.execute('''
                INSERT INTO activities (employee_id, activity_name, activity_category, 
                                      entries_awarded, awarded_by, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (employee_id, 'Points Reset', 'system', -old_total, 
                 request.current_user['user_id'], f'Reset from {old_total} to 0'))
            
            conn.commit()
            
            # Log the action
            db.log_audit(
                request.current_user['user_id'],
                "Reset employee points",
                "employees",
                employee_id,
                old_values={'total_entries': old_total},
                new_values={'total_entries': 0},
                ip_address=get_remote_address()
            )
            
            return jsonify({
                'success': True, 
                'message': f'Reset {employee["name"]} points from {old_total} to 0'
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reset_all', methods=['POST'])
@login_required
@role_required('admin')
@limiter.limit("1 per hour")
def reset_all_data():
    """DANGEROUS: Reset all employee data - requires admin role and is rate limited"""
    try:
        data = request.get_json()
        confirmation = data.get('confirmation', '')
        
        if confirmation != 'RESET_ALL_DATA':
            return jsonify({'success': False, 'error': 'Invalid confirmation'}), 400
        
        with db.get_connection() as conn:
            # Create backup before reset
            backup_file = db.backup_database()
            
            # Mark all employees as inactive instead of deleting
            conn.execute('UPDATE employees SET is_active = 0')
            
            # Add system activity for the reset
            conn.execute('''
                INSERT INTO activities (employee_id, activity_name, activity_category, 
                                      entries_awarded, awarded_by, notes)
                SELECT id, 'System Reset', 'system', -total_entries, ?, 'All data reset'
                FROM employees WHERE total_entries > 0
            ''', (request.current_user['user_id'],))
            
            conn.commit()
            
            # Log the action
            db.log_audit(
                request.current_user['user_id'],
                "SYSTEM RESET - All employee data reset",
                "system",
                notes=f"Backup created: {backup_file}",
                ip_address=get_remote_address()
            )
            
            return jsonify({
                'success': True, 
                'message': 'All employee data has been reset',
                'backup_file': backup_file
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/import_excel', methods=['POST'])
@login_required
@role_required('manager')
@limiter.limit("5 per hour")
def import_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Invalid file type. Please upload .xlsx or .xls files only'}), 400
        
        # Check file size
        if len(file.read()) > app.config['MAX_FILE_SIZE']:
            return jsonify({'success': False, 'error': 'File too large'}), 400
        file.seek(0)  # Reset file pointer
        
        # Save uploaded file securely
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_PATH'], safe_filename)
        file.save(filepath)
        
        try:
            # Process the Excel file
            result = process_excel_file(filepath)
            
            if not result['success']:
                return jsonify({'success': False, 'error': f'Failed to process Excel file: {result["error"]}'}), 400
            
            # Import employees to database
            added_count = 0
            skipped_count = 0
            
            with db.get_connection() as conn:
                for employee_name in result['employees']:
                    # Check if employee already exists
                    cursor = conn.execute('SELECT id FROM employees WHERE name = ?', (employee_name,))
                    if cursor.fetchone():
                        skipped_count += 1
                        continue
                    
                    # Insert new employee
                    conn.execute('''
                        INSERT INTO employees (name, total_entries)
                        VALUES (?, ?)
                    ''', (employee_name, 0))
                    added_count += 1
                
                conn.commit()
            
            # Log the import
            db.log_audit(
                request.current_user['user_id'],
                f"Excel import: {added_count} employees added",
                "employees",
                new_values={
                    'filename': filename,
                    'added': added_count,
                    'skipped': skipped_count
                },
                ip_address=get_remote_address()
            )
            
            return jsonify({
                'success': True,
                'message': f'Successfully imported {added_count} new employees',
                'total_employees_found': len(result['employees']),
                'new_employees_added': added_count,
                'existing_employees_skipped': skipped_count,
                'file_info': {
                    'total_rows': result['total_rows'],
                    'columns': result['columns'],
                    'detected_name_column': result['detected_name_column']
                }
            })
            
        finally:
            # Clean up uploaded file
            if os.path.exists(filepath):
                os.remove(filepath)
        
    except Exception as e:
        return jsonify({'success': False, 'error': f'An error occurred: {str(e)}'}), 500

# New professional endpoints
@app.route('/api/raffle/conduct', methods=['POST'])
@login_required
@role_required('manager')
def conduct_raffle():
    """Conduct a raffle and record the winner"""
    try:
        data = request.get_json()
        prize = data.get('prize', 'Quarterly Prize')
        
        with db.get_connection() as conn:
            # Get all eligible employees (with entries > 0)
            cursor = conn.execute('''
                SELECT id, name, total_entries 
                FROM employees 
                WHERE is_active = 1 AND total_entries > 0
                ORDER BY name
            ''')
            
            eligible_employees = cursor.fetchall()
            
            if not eligible_employees:
                return jsonify({'success': False, 'error': 'No eligible employees found'}), 400
            
            # Calculate total entries and chances
            total_entries = sum(emp['total_entries'] for emp in eligible_employees)
            participants = []
            
            for emp in eligible_employees:
                chance = (emp['total_entries'] / total_entries) * 100
                participants.append({
                    'id': emp['id'],
                    'name': emp['name'],
                    'entries': emp['total_entries'],
                    'chance': round(chance, 2)
                })
            
            # For now, return the data for client-side selection
            # In a real implementation, you might want to do server-side selection
            return jsonify({
                'success': True,
                'participants': participants,
                'total_entries': total_entries,
                'total_participants': len(participants)
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/raffle/record_winner', methods=['POST'])
@login_required
@role_required('manager')
def record_raffle_winner():
    """Record the winner of a raffle"""
    try:
        data = request.get_json()
        winner_id = data.get('winner_id')
        prize = data.get('prize', 'Quarterly Prize')
        total_participants = data.get('total_participants', 0)
        total_entries = data.get('total_entries', 0)
        winning_chance = data.get('winning_chance', 0)
        
        if not winner_id:
            return jsonify({'success': False, 'error': 'Winner ID is required'}), 400
        
        with db.get_connection() as conn:
            # Verify winner exists
            cursor = conn.execute('SELECT name FROM employees WHERE id = ? AND is_active = 1', (winner_id,))
            winner = cursor.fetchone()
            
            if not winner:
                return jsonify({'success': False, 'error': 'Invalid winner ID'}), 400
            
            # Record raffle result
            cursor = conn.execute('''
                INSERT INTO raffle_history 
                (winner_id, prize, total_participants, total_entries, winning_chance, conducted_by)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (winner_id, prize, total_participants, total_entries, winning_chance, 
                 request.current_user['user_id']))
            
            raffle_id = cursor.lastrowid
            conn.commit()
            
            # Log the raffle
            db.log_audit(
                request.current_user['user_id'],
                f"Conducted raffle - Winner: {winner['name']}",
                "raffle_history",
                raffle_id,
                new_values={
                    'winner': winner['name'],
                    'prize': prize,
                    'participants': total_participants
                },
                ip_address=get_remote_address()
            )
            
            return jsonify({
                'success': True,
                'message': f'Raffle completed! Winner: {winner["name"]}',
                'winner_name': winner['name'],
                'raffle_id': raffle_id
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/analytics/dashboard', methods=['GET'])
@login_required
def analytics_dashboard():
    """Get analytics data for dashboard"""
    try:
        with db.get_connection() as conn:
            # Total employees
            cursor = conn.execute('SELECT COUNT(*) as total FROM employees WHERE is_active = 1')
            total_employees = cursor.fetchone()['total']
            
            # Total entries
            cursor = conn.execute('SELECT SUM(total_entries) as total FROM employees WHERE is_active = 1')
            total_entries = cursor.fetchone()['total'] or 0
            
            # Recent activities
            cursor = conn.execute('''
                SELECT a.activity_name, a.entries_awarded, a.created_at, e.name as employee_name
                FROM activities a
                JOIN employees e ON a.employee_id = e.id
                WHERE e.is_active = 1
                ORDER BY a.created_at DESC
                LIMIT 10
            ''')
            recent_activities = [dict(row) for row in cursor.fetchall()]
            
            # Top performers
            cursor = conn.execute('''
                SELECT name, total_entries, department
                FROM employees 
                WHERE is_active = 1 AND total_entries > 0
                ORDER BY total_entries DESC
                LIMIT 10
            ''')
            top_performers = [dict(row) for row in cursor.fetchall()]
            
            # Department breakdown
            cursor = conn.execute('''
                SELECT department, COUNT(*) as employee_count, SUM(total_entries) as total_entries
                FROM employees 
                WHERE is_active = 1 AND department IS NOT NULL
                GROUP BY department
                ORDER BY total_entries DESC
            ''')
            department_stats = [dict(row) for row in cursor.fetchall()]
            
            return jsonify({
                'success': True,
                'analytics': {
                    'total_employees': total_employees,
                    'total_entries': total_entries,
                    'recent_activities': recent_activities,
                    'top_performers': top_performers,
                    'department_stats': department_stats
                }
            })
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/backup', methods=['POST'])
@login_required
@role_required('admin')
def create_backup():
    """Create a database backup"""
    try:
        backup_file = db.backup_database()
        
        # Log the backup
        db.log_audit(
            request.current_user['user_id'],
            "Database backup created",
            notes=f"Backup file: {backup_file}",
            ip_address=get_remote_address()
        )
        
        return jsonify({
            'success': True,
            'message': 'Backup created successfully',
            'backup_file': os.path.basename(backup_file)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('NODE_ENV', 'development') == 'development'
    app.run(debug=debug, host='0.0.0.0', port=port)