#!/usr/bin/env python3
"""
Test script to debug Excel import issues
"""
import os
import sys
from openpyxl import load_workbook

def test_excel_file(filepath):
    """Test if we can read the Excel file"""
    print(f"Testing Excel file: {filepath}")
    
    # Check if file exists
    if not os.path.exists(filepath):
        print(f"ERROR: File does not exist: {filepath}")
        return False
    
    print(f"File exists: {os.path.getsize(filepath)} bytes")
    
    try:
        # Try to load the workbook
        print("Loading workbook...")
        workbook = load_workbook(filepath)
        print(f"Workbook loaded successfully")
        print(f"Sheet names: {workbook.sheetnames}")
        
        # Get the active sheet
        sheet = workbook.active
        print(f"Active sheet: {sheet.title}")
        print(f"Max row: {sheet.max_row}")
        print(f"Max column: {sheet.max_column}")
        
        # Read headers (first row)
        print("\nHeaders (first row):")
        headers = []
        for i, cell in enumerate(sheet[1], 1):
            header = str(cell.value) if cell.value else f"Column_{i}"
            headers.append(header)
            print(f"  Column {i}: {header}")
        
        # Show first few data rows
        print(f"\nFirst 3 data rows:")
        for row_num in range(2, min(6, sheet.max_row + 1)):
            row_data = []
            for cell in sheet[row_num]:
                value = str(cell.value) if cell.value is not None else ""
                row_data.append(value)
            print(f"  Row {row_num}: {row_data}")
        
        # Look for name-like columns
        print(f"\nLooking for name columns...")
        name_patterns = ['name', 'caregiver', 'employee', 'staff', 'firstname', 'lastname', 'full']
        
        potential_name_columns = []
        for i, header in enumerate(headers):
            header_lower = header.lower()
            for pattern in name_patterns:
                if pattern in header_lower:
                    potential_name_columns.append((i, header))
                    break
        
        if potential_name_columns:
            print("Found potential name columns:")
            for col_index, col_name in potential_name_columns:
                print(f"  Column {col_index + 1}: {col_name}")
                # Show some sample data from this column
                sample_data = []
                for row_num in range(2, min(6, sheet.max_row + 1)):
                    value = sheet.cell(row=row_num, column=col_index + 1).value
                    if value:
                        sample_data.append(str(value))
                print(f"    Sample data: {sample_data}")
        else:
            print("No obvious name columns found. Checking first column...")
            # Show data from first column
            sample_data = []
            for row_num in range(2, min(6, sheet.max_row + 1)):
                value = sheet.cell(row=row_num, column=1).value
                if value:
                    sample_data.append(str(value))
            print(f"    First column sample data: {sample_data}")
        
        return True
        
    except Exception as e:
        print(f"ERROR: Failed to process Excel file: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    filepath = r"C:\Users\bir_k\Downloads\caregivers_08-13-2025-12-13-07.xlsx"
    success = test_excel_file(filepath)
    
    if success:
        print("\n✓ Excel file can be read successfully")
    else:
        print("\n✗ Excel file processing failed")