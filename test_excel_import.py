#!/usr/bin/env python3
"""
Test the updated Excel import functionality
"""
import os
import sys
from openpyxl import load_workbook

def process_excel_file(filepath):
    """
    Process Excel file and extract employee names using openpyxl.
    This function will try to find employee names in common column patterns.
    """
    try:
        # Load the Excel file
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        # Get all data from the sheet
        data = []
        headers = []
        
        # Read headers (first row)
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value else "")
        
        # Read all data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append([str(cell) if cell is not None else "" for cell in row])
        
        employees = []
        first_name_col = None
        last_name_col = None
        full_name_col = None
        detected_columns = []
        
        # Look for first name and last name columns
        for i, header in enumerate(headers):
            header_lower = header.lower().strip()
            if 'first' in header_lower and 'name' in header_lower:
                first_name_col = i
                detected_columns.append(f"First Name (Column {i+1})")
            elif 'last' in header_lower and 'name' in header_lower:
                last_name_col = i
                detected_columns.append(f"Last Name (Column {i+1})")
            elif ('full' in header_lower or 'employee' in header_lower or 'caregiver' in header_lower or 'staff' in header_lower) and 'name' in header_lower:
                full_name_col = i
                detected_columns.append(f"Full Name (Column {i+1})")
        
        # If we found both first and last name columns, combine them
        if first_name_col is not None and last_name_col is not None:
            for row in data:
                if (len(row) > max(first_name_col, last_name_col) and 
                    row[first_name_col] and row[last_name_col]):
                    
                    first_name = str(row[first_name_col]).strip()
                    last_name = str(row[last_name_col]).strip()
                    
                    if (first_name and last_name and 
                        first_name.lower() not in ['none', 'null', 'nan', ''] and
                        last_name.lower() not in ['none', 'null', 'nan', '']):
                        
                        full_name = f"{first_name} {last_name}"
                        employees.append(full_name)
        
        # If we found a full name column, use that
        elif full_name_col is not None:
            for row in data:
                if len(row) > full_name_col and row[full_name_col]:
                    name = str(row[full_name_col]).strip()
                    if (len(name) > 2 and len(name) < 100 and 
                        any(c.isalpha() for c in name) and 
                        name.lower() not in ['none', 'null', 'nan', '']):
                        employees.append(name)
        
        # If no obvious name columns found, check first few columns
        elif not detected_columns:
            # Common column name patterns that might contain employee names
            name_patterns = ['name', 'caregiver', 'employee', 'staff']
            
            for i, header in enumerate(headers):
                if header and any(pattern in header.lower() for pattern in name_patterns):
                    full_name_col = i
                    detected_columns.append(f"Name (Column {i+1}: {header})")
                    break
            
            if full_name_col is not None:
                for row in data:
                    if len(row) > full_name_col and row[full_name_col]:
                        name = str(row[full_name_col]).strip()
                        if (len(name) > 2 and len(name) < 100 and 
                            any(c.isalpha() for c in name) and 
                            name.lower() not in ['none', 'null', 'nan', '']):
                            employees.append(name)
            else:
                # Last resort: check first column for names
                detected_columns.append("First Column (assumed names)")
                for row in data:
                    if len(row) > 0 and row[0]:
                        potential_name = str(row[0]).strip()
                        if (len(potential_name) > 2 and len(potential_name) < 100 and 
                            any(c.isalpha() for c in potential_name) and
                            potential_name.lower() not in ['none', 'null', 'nan', '']):
                            employees.append(potential_name)
        
        if workbook:
            workbook.close()
        
        return {
            'success': True,
            'employees': list(set(employees)),  # Remove duplicates
            'total_rows': len(data),
            'columns': headers,
            'detected_columns': detected_columns,
            'employees_found': len(set(employees))
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

if __name__ == "__main__":
    filepath = r"C:\Users\bir_k\Downloads\caregivers_08-13-2025-12-13-07.xlsx"
    print(f"Testing Excel import with: {filepath}")
    
    result = process_excel_file(filepath)
    
    if result['success']:
        print(f"SUCCESS!")
        print(f"Total rows processed: {result['total_rows']}")
        print(f"Columns found: {result['columns']}")
        print(f"Detected name columns: {result['detected_columns']}")
        print(f"Employees found: {result['employees_found']}")
        print("\nFirst 10 employees:")
        for i, employee in enumerate(result['employees'][:10]):
            print(f"  {i+1}. {employee}")
        
        if len(result['employees']) > 10:
            print(f"... and {len(result['employees']) - 10} more")
            
    else:
        print(f"FAILED: {result['error']}")